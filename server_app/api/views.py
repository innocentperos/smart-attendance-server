import datetime
import os
from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework.decorators import action, api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework import status
from server_app.api.decorators import lecturer_only
from server_app.api.forms import AttendanceForm, CourseForm, LoginForm, StudentForm
from server_app.api.serializers import AttendanceSerializer, CourseSerializer, StudentAttendanceSerializer
from server_app.models import Attendance, Course, Student, StudentAttendance, UserType, XSession, random_id
from rest_framework import viewsets
from django.shortcuts import get_object_or_404
from server_app.models import Lecturer
from rest_framework import exceptions
import face_recognition
import openpyxl as excel
from django.conf import settings
from pathlib import Path

TOLERANCE = 0.42


@api_view(['post'])
@permission_classes([AllowAny])
def login(request):
    form = LoginForm(request.data)

    if not form.is_valid():

        return Response(
            status=status.HTTP_406_NOT_ACCEPTABLE,
            data=form.errors
        )

    data = form.cleaned_data

    try:
        user = get_user_model().objects.get(username=data['email'])

        if not user.check_password(data['password']):
            return Response(
                status=status.HTTP_401_UNAUTHORIZED,
                data={
                    "message": "Wrong Password"
                }
            )

        user_type = UserType.objects.get(user=user)

        session_id = random_id(12)
        session = XSession(user_type=user_type)

        session.session = session_id
        session.expires = datetime.datetime.now() + datetime.timedelta(days=7)

        session.save()

        return Response(
            {
                "session": session.session
            }
        )

    except get_user_model().DoesNotExist:
        return Response(
            status=status.HTTP_401_UNAUTHORIZED,
            data={
                "message": "Wrong password or email address"
            }
        )
    except UserType.DoesNotExist:
        if data['email'] == 'admin':
            user_type = UserType(user=user, user_type="ADMIN").save()
        return Response(
            status=status.HTTP_401_UNAUTHORIZED,
            data={
                "message": "Invalid user type"
            }
        )

    pass


@api_view(['get'])
def user(request):
    user = request.user

    user_type = UserType.objects.filter(user=user).first()

    return Response(
        {
            'type': user_type.user_type
        }
    )


class LecturerViewSet(viewsets.ViewSet):

    queryset = Lecturer.objects.all()

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def signup(self, request, *args, **kwargs):
        form = LoginForm(request.data)

        if not form.is_valid():
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data=form.errors
            )
        User = get_user_model()
        data = form.cleaned_data
        try:
            user = User.objects.get(username=data['email'])

            return Response(
                status=status.HTTP_409_CONFLICT,
                data={
                    "message": "A user with thesame email already exists"
                }
            )
        except User.DoesNotExist:

            email = data['email']
            user = User(email=email, username=email)
            user.set_password(data['password'])

            user.save()

            user_type = UserType(user=user, user_type="LECTURER")

            user_type.save()
            lecturer = Lecturer(email=email)
            lecturer.user = user
            lecturer.save()
            session_id = random_id()

            session = XSession()
            session.user_type = user_type
            session.session = session_id

            session.save()

            return Response(
                {
                    "session": session.session
                }
            )


class CourseViewSet(viewsets.ViewSet):
    queryset = Course.objects.all()

    def list(self, request):
        user = request.user
        try:
            user_type = UserType.objects.get(user=user)
            if user_type.user_type != "LECTURER":
                raise exceptions.AuthenticationFailed(
                    "Please login with a lecturer account")
        except UserType.DoesNotExist:
            raise exceptions.AuthenticationFailed(
                "Please login with a valid account")

        return Response(CourseSerializer(self.queryset.filter(lecturer=Lecturer.objects.get(user=user)), many=True).data)

    def create(self, request):
        user = request.user

        try:
            user_type = UserType.objects.get(user=user)
        except UserType.DoesNotExist:
            raise exceptions.AuthenticationFailed(
                "Please Login in with a lecturer account")

        print(user_type.user_type)

        if user_type.user_type != "LECTURER":
            raise exceptions.AuthenticationFailed(
                "2 Please Login in with a lecturer account")

        form = CourseForm(request.data)

        if not form.is_valid():
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data=form.errors
            )
        data = form.cleaned_data

        try:
            lecturer = Lecturer.objects.get(user=user)
        except Lecturer.DoesNotExist:
            raise exceptions.AuthenticationFailed(
                "Please login with a valid lecturer account")

        try:
            course = self.queryset.get(lecturer=lecturer, code=data['code'])
            return Response(
                status=status.HTTP_409_CONFLICT,
                data={"message": f"Course {course.title} with thesame code exist"}
            )
        except Course.DoesNotExist:
            pass

        course = form.save(commit=False)
        course.lecturer = lecturer
        course.save()

        return Response(
            CourseSerializer(course).data
        )


class AttendanceViewSet(viewsets.ViewSet):
    queryset = Attendance.objects.all()

    @lecturer_only
    def list(self, request):
        course_id = request.query_params.get("course", None)

        if not course_id:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": "missing course parameter"
                }
            )
        lecturer = request.lecturer
        try:
            course = Course.objects.get(pk=course_id)
        except Course.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Course not found"
                }
            )

        if course.lecturer != lecturer:
            return Response(
                status=status.HTTP_401_UNAUTHORIZED,
                data={
                    "message": "You cant view courses attendances that are not yours"
                }
            )
        sets = self.queryset.filter(course=course)

        return Response(
            AttendanceSerializer(sets, many=True).data
        )

    @lecturer_only
    def create(self, request):

        form = AttendanceForm(request.data)

        if not form.is_valid():
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data=form.errors
            )

        code = random_id(6)

        attendance = form.save(commit=False)
        attendance.code = code
        attendance.is_open = True
        attendance.commit = False
        attendance.datetime = datetime.datetime.now()

        try:
            attendance.save()
            return Response(
                status=status.HTTP_201_CREATED,
                data=AttendanceSerializer(attendance).data
            )
        except Exception as error:
            return Response(
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                data={
                    "message": "Something went wrong",
                    "error": str(error)
                }
            )

    @lecturer_only
    def detail(self, request, pk):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Attendance not found"
                }
            )

        attendances = StudentAttendance.objects.filter(attendance=attendance)

        return Response(
            StudentAttendanceSerializer(attendances, many=True).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def commit(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Attendance not found"
                }
            )

        attendance.commit = True
        attendance.is_open = False
        attendance.save()

        return Response(
            AttendanceSerializer(attendance).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def close(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Attendance not found"
                }
            )
        if not attendance.commit:
            attendance.is_open = False
            attendance.save()

        return Response(
            AttendanceSerializer(attendance).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def open(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Attendance not found"
                }
            )

        attendance.is_open = True
        attendance.save()

        return Response(
            AttendanceSerializer(attendance).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def delete(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Attendance not found"
                }
            )

        attendance.delete()

        return Response(
            AttendanceSerializer(attendance).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def attendances(self, request, pk=None):
        try:
            course = Course.objects.get(pk=pk)
        except Course.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Course could not be found"
                }
            )

        attendances = self.queryset.filter(course=course)

        return Response(
            AttendanceSerializer(attendances, many=True).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def attendance_list(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Course Attendance list could not be found"
                }
            )

        attendances = StudentAttendance.objects.filter(attendance=attendance)

        return Response(
            StudentAttendanceSerializer(attendances, many=True).data
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def export_all(self, request, pk=None):
        mark = int(request.query_params.get("mark",1))
        try:
            course = Course.objects.get(pk=pk)
        except Course.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Course could not be found"
                }
            )

        course_attendances = self.queryset.filter(course=course)

        students_attendances = StudentAttendance.objects.filter(
            attendance__course=course)

        students = {}
        index = 1

        for student_attendance in students_attendances:
            student = student_attendance.student

            if (student.id) not in students:
                students[(student.id)] = {
                    "index": index + 1,
                    "matric_number": student.matric_number,
                    "attendances": 0
                }
                index = index + 1

        # workbook = excel.Workbook()
        sheet = {}

        sheet['A1'] = "Student Matric Number"
        sheet['B1'] = "Date"

        for (index, attendance) in enumerate(course_attendances):
            

            students_attendances = StudentAttendance.objects.filter(
                attendance=attendance)
           

            sheet[f'{chr(ord("B")+index)}1'] = attendance.date().split(";")[0]

            for student_attendance in students_attendances:
                student_sheet = students[student_attendance.student.id]
                
                sheet[f'A{student_sheet["index"]}'] = student_sheet["matric_number"]
                sheet[f'{chr(ord("B")+index)}{student_sheet["index"]}'] = 'Present'
                student_sheet["attendances"] = int(student_sheet["attendances"])+1

        sheet_two={}
        #  workbook.create_sheet("Students Scores")
        index = 2
        sheet_two['A1'] = "Matric Number"
        sheet_two['B1'] ="Presents"
        sheet_two['C1'] = "Absent"
        sheet_two['D1'] = "Total"
        sheet_two['E1'] = "Mark"


        for id in students:
            student = students[id]
            try:
                sheet_two[f'A{index}'] =student['matric_number']
            except Exception as error:
                print(student)
                print(error)
            
            sheet_two[f'B{index}'] =student['attendances']
            sheet_two[f'C{index}'] = int(course_attendances.count()) - int(student['attendances'])
            sheet_two[f'D{index}'] = int(course_attendances.count()) 
            sheet_two[f'E{index}'] = float(int(student['attendances']) / int(course_attendances.count())) * mark

            index = index +1

        # filename = f"{attendance.course.code}-{attendance.pk}-attendance-ALL.xlsx"
        # path = Path.joinpath(settings.MEDIA_ROOT, "exports", filename)
        # workbook.save(filename=path)

        # stt = (settings.MEDIA_URL+"exports/" + filename)


        return Response(
            {
                "attendance_sheet":sheet,
                "score_sheet" : sheet_two
            }
        )

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def course_attendance(self, request, pk = None):
        course = get_object_or_404(Course.objects.all(), pk)

        attendances = StudentAttendance.objects.filter(attendance__course = course)

        return Response(StudentAttendanceSerializer(attendances, many=True).data)

    @action(detail=True, methods=['get'], permission_classes=[AllowAny])
    @lecturer_only
    def export(self, request, pk=None):
        try:
            attendance = self.queryset.get(pk=pk)
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Course Attendance list could not be found"
                }
            )

        attendances = StudentAttendance.objects.filter(attendance=attendance)

        workbook = excel.Workbook()

        sheet = workbook.active

        sheet['A1'] = "Student Matric Number"

        for (index, item) in enumerate(attendances):
            sheet[f'A{index+2}'] = item.student.matric_number

        filename = f"attendance.xlsx"
        path = os.path.join(settings.MEDIA_ROOT, filename)
        # path = Path.joinpath(settings.MEDIA_ROOT, "exports", filename)
        print(path)
        file = open(path,'w');
        file.close()
        workbook.save(filename=path)

        stt = (settings.MEDIA_URL+"/" + filename)
        return Response(
            {
                "url": stt
            }
        )

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def validate(self, request):
        code = request.data.get("attendance_code", None)
        matric_number = request.data.get("matric_number", None)

        try:
            attendance = Attendance.objects.get(code__iexact=code)

            if not attendance.is_open:
                return Response(
                    status=status.HTTP_405_METHOD_NOT_ALLOWED,
                    data={
                        "message": "Attendance list found but is not open"
                    }
                )
            elif attendance.commit:
                return Response(
                    status=status.HTTP_405_METHOD_NOT_ALLOWED,
                    data={
                        "message": "Attendance list already submitted"
                    }
                )
        except Attendance.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "No open attendance found, make sure the attendance code is correct"
                }
            )

        try:
            student = Student.objects.get(matric_number=matric_number)

            return Response(
                {
                    "course": CourseSerializer(attendance.course).data,
                    "date": attendance.datetime,
                    'lat': attendance.lat,
                    'long': attendance.long
                }
            )
        except Student.DoesNotExist:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "message": "Student not enrolled"
                }
            )

    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def attend(self, request):
        code = request.data.get("attendance_code", None)

        if not code:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Please provide the course Attendance code"}
            )

        matric_number = request.data.get("matric_number", None)
        if not matric_number:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Please provide your matric number"}
            )
        distance = request.data.get("distance", None)
        if not distance:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Sorry could not determine your position, please make sure that GPS is enabled and permission granted to the app"}

            )

        lat = request.data.get("lat", None)
        if not lat:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Sorry could not determine your position, please make sure that GPS is enabled and permission granted to the app"}
            )

        long = request.data.get("long", None)
        if not long:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Sorry could not determine your position, please make sure that GPS is enabled and permission granted to the app"}

            )

        image = request.FILES.get("image")
        if not image:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={"message": "Please Capture an image showing your face"}
            )

        attendance = self.queryset.filter(code__iexact=code).first()

        if not attendance:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": "No attendance list was found, make sure you provide a valid attendance code"}
            )

        if not attendance.is_open:
            return Response(
                status=status.HTTP_405_METHOD_NOT_ALLOWED,
                data={
                    "message": "Attendance list found but is not open"
                }
            )
        elif attendance.commit:
            return Response(
                status=status.HTTP_405_METHOD_NOT_ALLOWED,
                data={
                    "message": "Attendance list already submitted"
                }
            )

        student = Student.objects.filter(matric_number=matric_number).first()
        if not student:
            return Response(
                status=status.HTTP_405_METHOD_NOT_ALLOWED,
                data={
                    "message": "This student is not enrolled yet"
                }
            )

        try:
            student_attendance = StudentAttendance.objects.get(
                student=student, attendance=attendance)

            return Response(
                status=status.HTTP_405_METHOD_NOT_ALLOWED,
                data={
                    "message": "This student is already on the list"
                }
            )
        except StudentAttendance.DoesNotExist:
            pass

        student_attendance = StudentAttendance()
        student_attendance.student = student
        student_attendance.attendance = attendance
        student_attendance.attendance_image = image
        student_attendance.lat = lat
        student_attendance.long = long

        student_attendance.save()

        attendance_image = face_recognition.load_image_file(
            student_attendance.attendance_image.path)
        face_location_list = face_recognition.face_locations(attendance_image)

        if len(face_location_list) > 1:
            student_attendance.delete()
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": f"Sorry {len(face_location_list)} faces were detected, only one face should be in the image"
                }
            )
        if len(face_location_list) < 1:
            student_attendance.delete()
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": f"Sorry no face was detected, make sure your face is showing in the image"
                }
            )

        image_of_student = face_recognition.load_image_file(student.image.path)
        student_face_encoding = face_recognition.face_encodings(image_of_student)[
            0]

        image_of_attendance = face_recognition.load_image_file(
            student_attendance.attendance_image.path)
        attendance_face_encoding = face_recognition.face_encodings(image_of_attendance)[
            0]

        result = face_recognition.compare_faces(
            [student_face_encoding], attendance_face_encoding, TOLERANCE)

        print("Call response")
        print(result)

        if result[0]:
            student_attendance.commit = True
            student_attendance.datetime = timezone.now()
            student_attendance.save()

            return Response(
                {
                    "datetime": student_attendance.datetime
                }
            )
        else:
            student_attendance.delete()

            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": f"Sorry your face did not match with the one you submitted during enrollment"
                }
            )


class StudentViewSet(viewsets.ViewSet):
    queryset = Student.objects.all()
    permission_classes = [AllowAny]

    def create(self, request):

        matric_number = request.data.get("matric_number", None)
        image = request.FILES.get("image")

        if not image or not matric_number:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": "Please provide both matric number and image"
                }
            )

        try:
            student = self.queryset.get(matric_number=matric_number)

            return Response(
                status=status.HTTP_409_CONFLICT,
                data={
                    "image": student.image.path,
                    "image_url": student.image.url,
                    "message": "Looks like this student matric number has been enrolled"
                }
            )
        except Student.DoesNotExist:
            student = Student(image=image, matric_number=matric_number)
            student.save()

            image = face_recognition.load_image_file(student.image.path)
            face_location_list = face_recognition.face_locations(image)

            if len(face_location_list) > 1:
                student.delete()
                return Response(
                    status=status.HTTP_406_NOT_ACCEPTABLE,
                    data={
                        "message": f"Sorry {len(face_location_list)} faces were detected, only one face should be in the image"
                    }
                )
            if len(face_location_list) < 1:
                student.delete()
                return Response(
                    status=status.HTTP_406_NOT_ACCEPTABLE,
                    data={
                        "message": f"Sorry no face was detected, make sure your face is showing in the image"
                    }
                )

            return Response(
                status=status.HTTP_201_CREATED,
                data={
                    "matric_number": f"{matric_number}"
                }
            )

    @action(detail=False, methods=['get'], permission_classes=[AllowAny])
    def check(self, request):
        matric_number = request.query_params.get("matric_number", None)

        if not matric_number:
            return Response(
                status=status.HTTP_406_NOT_ACCEPTABLE,
                data={
                    "message": "Please provide student matric number"
                }
            )

        student = self.queryset.filter(
            matric_number__iexact=matric_number).first()

        if student:
            return Response({
                "matric_number": matric_number
            })
        else:
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={
                    "matric_number": ""
                }
            )
